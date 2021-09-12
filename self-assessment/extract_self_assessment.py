#!/usr/bin/env python3

import sys
import os
import re
import openpyxl
import logging


logging.basicConfig(level=logging.DEBUG)


def parse_formula(ws, formula):
    logging.debug("formula: {}".format(formula))
    if not "=" in str(formula):
        if formula:
            try:
                return float(formula)
            except:
                return 0
        return 0
    logging.debug("here")
    formula = formula.replace(' ', '')
    formula = formula.strip("=")
    logging.debug("here2")
    if 'SUM' in formula:
        if ':' in formula:
            row_start = int(formula[4:-1].split(':')[0][1:])
            row_end = int(formula[4:-1].split(':')[1][1:])
            col = formula[4:-1].split(':')[0][0:1]
        else:
            row_start = int(formula[5:-1])
            row_end = row_start
            col = formula[4:1]
        value = 0
        logging.debug("col: {}, row_start: {}, row_end: {}".format(col, row_start, row_end))
        for idx in range(row_start, row_end+1):
            value += parse_formula(ws, '={}{}'.format(col, idx))
        return value
    if re.search("[a-zA-Z]", formula):
        if '+' in formula:
            value = 0
            for s in formula.split('+'):
                value += parse_formula(ws, '=' + s)
            return value
        if '*' in formula:
            value = 0
            for s in formula.split('*'):
                value *= parse_formula(ws, '=' + s)
            return value
        logging.debug("formula: {}".format(formula))
        logging.debug("cell: {}".format(ws[formula].value))
        return parse_formula(ws, ws[formula].value)
    logging.debug("here3")
    if re.search("[\+\*\/\-]", formula):
        return float(eval(formula))
    logging.debug("here4")
    return float(formula)


def process_file(filename):
    wb = openpyxl.load_workbook(filename)
    ws = wb.active

    logging.debug("ws['A7']: {}".format(ws['A7'].value))
    logging.debug("ws['A10']: {}".format(ws['A10'].value))

    if "renume" in str(ws['A7'].value):
        logging.debug("first type")
        name = ws['A7'].value.split(':')[1].strip()
        position = ws['A8'].value.split(':')[1].strip()
        education = parse_formula(ws, ws['C10'].value)
        research = parse_formula(ws, ws['C11'].value)
        manage = parse_formula(ws, ws['C12'].value)
        prestige = parse_formula(ws, ws['C13'].value)
        community = parse_formula(ws, ws['C14'].value)
        others = parse_formula(ws, ws['C15'].value)
    elif "renume" in str(ws['A10'].value):
        logging.debug("second type")
        name = ws['E10'].value
        position = ws['E11'].value
        education = parse_formula(ws, ws['H16'].value)
        research = parse_formula(ws, ws['H17'].value)
        manage = parse_formula(ws, ws['H18'].value)
        prestige = parse_formula(ws, ws['H19'].value)
        community = parse_formula(ws, ws['H20'].value)
        others = parse_formula(ws, ws['H21'].value)
    elif "renume" in str(ws['A8'].value):
        logging.debug("second type")
        name = ws['E8'].value
        position = ws['E9'].value
        education = parse_formula(ws, ws['H11'].value)
        research = parse_formula(ws, ws['H12'].value)
        manage = parse_formula(ws, ws['H13'].value)
        prestige = parse_formula(ws, ws['H14'].value)
        community = parse_formula(ws, ws['H15'].value)
        others = parse_formula(ws, ws['H16'].value)

    result = {
            "name": name,
            "position": position,
            "education": education,
            "research": research,
            "manage": manage,
            "prestige": prestige,
            "community": community,
            "others": others
            }
    logging.debug(result)

    return result


def main():
    if len(sys.argv) != 2:
        print("Usage: {} file".format(sys.argv[0]), file=sys.stderr)
        sys.exit(1)

    f = sys.argv[1]
    logging.info("processing {}".format(f))
    print("Nume\tPoziție\tEducație\tCercetare\tManagement\tRecunoaștere\tComunitate\tAltele")
    try:
        ret = process_file(f)
        print("{}\t{}\t{}\t{}\t{}\t{}\t{}\t{}".format(
            ret["name"],
            ret["position"],
            int(ret["education"]),
            int(ret["research"]),
            int(ret["manage"]),
            int(ret["prestige"]),
            int(ret["community"]),
            int(ret["others"])))
    except Exception as e:
        logging.error("Error processing {}: {}".format(f, str(e)))


if __name__ == "__main__":
    sys.exit(main())
